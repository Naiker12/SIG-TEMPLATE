
import os
import uuid
import pandas as pd
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from fastapi import HTTPException, UploadFile, status
from copy import copy
from fastapi.responses import FileResponse
from app import schemas

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


def duplicate_rows_with_format(ws: Worksheet, cantidad_col: str = "cantidad"):
    """
    Duplicates rows in a worksheet based on a 'cantidad' column, preserving cell formatting.
    If the 'cantidad' column doesn't exist, it does nothing.
    """
    try:
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        headers_lower = [str(h).strip().lower() if h else "" for h in headers]

        if cantidad_col not in headers_lower:
            # If the column doesn't exist, just return without doing anything.
            return

        col_index = headers_lower.index(cantidad_col) + 1
        data_to_duplicate = []

        # Read all data and styles first
        for row_idx in range(2, ws.max_row + 1):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append({
                    "value": cell.value,
                    "style": copy(cell._style) if cell.has_style else None
                })
            
            cantidad_val = ws.cell(row=row_idx, column=col_index).value
            try:
                # Ensure that None or non-numeric values default to 1
                count = int(cantidad_val) if pd.notna(cantidad_val) else 1
            except (ValueError, TypeError):
                count = 1
            count = max(1, count)
            data_to_duplicate.append((row_data, count))

        # Delete old rows (from bottom to top to avoid shifting issues if needed)
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)

        # Write duplicated data
        current_row_to_write = 2
        for row_data, count in data_to_duplicate:
            for _ in range(count):
                for col_idx, cell_data in enumerate(row_data, start=1):
                    new_cell = ws.cell(row=current_row_to_write, column=col_idx, value=cell_data["value"])
                    if cell_data["style"]:
                        new_cell._style = copy(cell_data["style"])
                current_row_to_write += 1
    except Exception as e:
        # Provide a more specific error message if possible
        raise HTTPException(status_code=500, detail=f"Error al duplicar filas en Excel: {e}")


async def upload_excel(file: UploadFile):
    """
    Handles the upload and processing of an Excel file.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser de tipo Excel (.xlsx, .xls)")

    file_id = f"{uuid.uuid4()}_{file.filename}"
    file_path = os.path.join(UPLOAD_DIR, file_id)

    # Save the uploaded file
    try:
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"No se pudo guardar el archivo: {e}")
    
    # Process the saved file
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        duplicate_rows_with_format(ws, "cantidad")
        wb.save(file_path)
    except HTTPException as e:
        # Re-raise HTTP exceptions from duplicate_rows_with_format
        raise e
    except Exception as e:
        # Clean up the file if processing fails and raise a generic error
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo Excel: {e}")

    return {"message": "Archivo subido y procesado correctamente", "file_id": file_id}


def preview_excel(file_id: str, page: int = 1, page_size: int = 10):
    """
    Provides a paginated JSON preview of an Excel file using openpyxl for robust reading.
    Adds a unique 'id' to each row for frontend identification.
    """
    file_path = os.path.join(UPLOAD_DIR, file_id)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    try:
        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb.active
        
        # Extract headers and create column definitions
        headers = [cell.value for cell in ws[1]]
        columns = [{"accessorKey": str(h), "header": str(h)} for h in headers if h is not None]
        
        # Extract rows as dictionaries
        all_rows = []
        # Add a unique ID to each row based on its physical row number
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {
                "id": row_idx,  # Unique ID based on row number
                **{headers[i]: value for i, value in enumerate(row)}
            }
            all_rows.append(row_dict)
            
        total_rows = len(all_rows)
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        paginated_data = all_rows[start_index:end_index]

        return {
            "page": page,
            "pageSize": page_size,
            "totalRows": total_rows,
            "totalPages": (total_rows + page_size - 1) // page_size,
            "columns": columns,
            "data": paginated_data,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al leer el archivo Excel: {e}")


def modify_excel(file_id: str, modifications: List[Dict[str, Any]]):
    """
    Modifies an Excel file based on provided data. (Placeholder)
    """
    file_path = os.path.join(UPLOAD_DIR, file_id)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    # This is a complex operation. The provided logic has issues.
    # For now, this is a placeholder.
    raise HTTPException(status_code=501, detail="La modificación de celdas aún no está implementada.")


def duplicate_row(file_id: str, row_id: int, count: int):
    """
    Duplicates a specific row in the Excel file `count` number of times.
    """
    file_path = os.path.join(UPLOAD_DIR, file_id)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    if count <= 0:
        return {"message": "No se realizaron duplicados."}

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        if not (2 <= row_id <= ws.max_row):
            raise HTTPException(status_code=404, detail=f"Fila con ID {row_id} no encontrada.")

        # Get the data and styles from the source row
        source_row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_id, column=col_idx)
            source_row_data.append({
                "value": cell.value,
                "style": copy(cell._style) if cell.has_style else None
            })

        # Insert new rows below the source row
        ws.insert_rows(row_id + 1, count)

        # Write the duplicated data into the new rows
        for i in range(count):
            target_row_idx = row_id + 1 + i
            for col_idx, cell_data in enumerate(source_row_data, start=1):
                new_cell = ws.cell(row=target_row_idx, column=col_idx, value=cell_data["value"])
                if cell_data["style"]:
                    new_cell._style = copy(cell_data["style"])
        
        wb.save(file_path)

        return {"message": f"{count} duplicado(s) de la fila {row_id} creado(s) con éxito."}

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al duplicar la fila: {e}")


def download_excel(file_id: str):
    """
    Allows downloading of the processed Excel file.
    """
    file_path = os.path.join(UPLOAD_DIR, file_id)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    # Extract original filename
    original_filename = "_".join(file_id.split("_")[1:])
    
    return FileResponse(file_path, filename=original_filename)
